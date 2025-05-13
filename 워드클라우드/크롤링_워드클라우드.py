from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, UnexpectedAlertPresentException, NoAlertPresentException
from selenium.webdriver.common.alert import Alert
from bs4 import BeautifulSoup
from datetime import datetime
import os
import time
from random import uniform
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook

# Firefox 옵션 설정
options = Options()
options.set_preference('general.useragent.override', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0')
options.set_preference("network.proxy.type", 0)

# Service 객체에 geckodriver 경로 지정
path = "/Users/user/Desktop/nh/geckodriver.exe"
service = Service(executable_path=path)

# Firefox 드라이버 시작
driver = webdriver.Firefox(service=service, options=options)

# 네이버 지식인 크롤링 함수들
def get_keyword(text):
    return text.replace(" ", "%20")

def sort_kind(index):
    if index == 1:
        return 'vcount'
    elif index == 2:
        return 'date'
    else:
        return 'none'

# 검색 키워드와 페이지 탐색 설정
keyword = 'etf 투자'
driver.get('https://kin.naver.com/search/list.nhn?query=' + get_keyword(keyword))
time.sleep(uniform(0.1, 1.0))

page_index = 1
f = '2024.01.01'
t = '2024.11.07'
period_txt = "&period=" + f + ".%7C" + t + "."
_sort_kind = sort_kind(2)
date = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')

# 결과 저장 디렉터리 생성
if not os.path.exists("result"):
    os.makedirs("result")

# URL 저장
with open(f"result/url_list_{keyword.replace(' ', '+')}_{date}.txt", 'w') as f:
    page_url = []
    while True:
        time.sleep(uniform(0.01, 1.0))
        driver.get('https://kin.naver.com/search/list.nhn?' + "&sort=" + _sort_kind + '&query=' + get_keyword(keyword) + period_txt + "&section=kin" + "&page=" + str(page_index))
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        tags = soup.find_all('a', class_="_nclicks:kin.txt _searchListTitleAnchor")
        for tag in tags:
            url = tag['href'].replace('amp;', '')
            page_url.append(url)
            f.write(url + "\n")

        try:
            post_number = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'number'))
            ).text
            post_number = post_number.replace("(", "").replace(")", "")
            current_number = post_number.split('/')[0].split('-')[1].replace(',', '')
            total_number = post_number.split('/')[1].replace(',', '')

            if int(current_number) == int(total_number):
                break
            else:
                page_index += 1

        except TimeoutException:
            print("Timeout: 'number' 클래스 요소를 찾는 데 시간이 초과되었습니다.")
            break

        except NoSuchElementException:
            print("NoSuchElementException: 'number' 클래스 요소를 찾을 수 없습니다.")
            break

# Excel 파일 작성
filename = f'result/{keyword.replace(" ", ".")}_{date}_crawling_result.xlsx'
wb = Workbook()
sheet = wb.active
sheet.append(['제목', '질문'])

# 제목과 질문에 색상 채우기
for j in range(1, 3):
    sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

count = 0

# 페이지 URL을 크롤링하며 제목과 질문 가져오기
for i in page_url:
    try:
        driver.get(i)
        title = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'title'))
        ).text

        try:
            question_txt = driver.find_element(By.CLASS_NAME, 'c-heading__content').text
        except NoSuchElementException:
            question_txt = ""

        sheet.append([title, question_txt])
        count += 1
        print(f"크롤링된 게시글 수: {count}")

    except UnexpectedAlertPresentException:
        try:
            alert = Alert(driver)
            alert.accept()
            print("팝업 창이 닫혔습니다.")
        except NoAlertPresentException:
            pass

    except TimeoutException:
        print(f"TimeoutException: 페이지 로드에 실패했습니다: {i}")

wb.save(filename)
print(f"크롤링 완료. 결과는 {filename}에 저장되었습니다.")

# 드라이버 종료
driver.quit()

import matplotlib.pyplot as plt
from collections import Counter
from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator
import numpy as np
from PIL import Image
import matplotlib.cm as cm

# 크롤링된 텍스트 데이터를 하나의 문자열로 합치기
all_text = ' '.join([title + ' ' + question_txt for title, question_txt in zip(titles, questions)])

# 텍스트 전처리 (특수 문자 제거 등 필요 시 추가)
import re
all_text = re.sub(r'[^가-힣a-zA-Z\s]', '', all_text)  # 한글과 영어 문자만 남기기

# 단어 빈도 계산
words = all_text.split()  # 공백 기준으로 단어 나누기
word_freq = Counter(words)  # 빈도 계산


def color_func(word, font_size, position, orientation, random_state=None, **kwargs):
    return "hsl({:d},{:d}%,{:d}%)".format(
        np.random.randint(212, 313),  # Hue 범위
        np.random.randint(26, 32),    # Saturation 범위
        np.random.randint(45, 80)     # Lightness 범위
    )


# 마스크 이미지 불러오기
mask_image = np.array(Image.open("C:/Users/user/Desktop/man6.png"))  # 마스크 이미지 파일 경로

colormap = plt.get_cmap('RdYlBu')

# 워드클라우드 생성
wordcloud = WordCloud(
    font_path='C:/Windows/Fonts/H2GTRE.TTF',  # 한글 폰트 파일 경로
    width=800,
    height=400,
    background_color='black',
    mask=mask_image,
    #contour_color='black',  # 윤곽선 색상
    #contour_width=1,        # 윤곽선 두께
    colormap=colormap
).generate_from_frequencies(word_freq)




# 결과 출력
plt.figure(figsize=(10, 10))
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis('off')  # 축 제거
plt.show()
